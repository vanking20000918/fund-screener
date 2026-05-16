"""
报告生成模块
1. HTML 邮件正文(Top N 推荐列表+评分+各指标明细)
2. Excel 附件(完整候选池详细数据)
"""
import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from .config import OUTPUT_DIR, TOP_N, SCORE_WEIGHTS

logger = logging.getLogger(__name__)


# ============================================================================
# HTML 邮件正文
# ============================================================================
HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
    body {{ font-family: -apple-system, "Microsoft YaHei", Arial; color: #333; line-height: 1.6; max-width: 900px; margin: 0 auto; padding: 20px; }}
    h1 {{ color: #1f4e78; border-bottom: 3px solid #2e75b6; padding-bottom: 10px; }}
    h2 {{ color: #1f4e78; margin-top: 30px; }}
    .summary-box {{ background: #e2efda; border-left: 4px solid #00b050; padding: 15px; margin: 20px 0; border-radius: 4px; }}
    .warning-box {{ background: #fff2cc; border-left: 4px solid #ffc000; padding: 12px; margin: 15px 0; border-radius: 4px; font-size: 13px; }}
    table {{ border-collapse: collapse; width: 100%; margin: 15px 0; font-size: 13px; }}
    th {{ background: #2e75b6; color: white; padding: 10px 8px; text-align: center; }}
    td {{ padding: 8px; border-bottom: 1px solid #d9d9d9; text-align: center; }}
    tr:nth-child(even) {{ background: #f8f9fa; }}
    .rank {{ font-weight: bold; color: #1f4e78; width: 40px; }}
    .grade-A {{ background: #00b050; color: white; padding: 3px 8px; border-radius: 3px; font-weight: bold; }}
    .grade-B {{ background: #92d050; color: white; padding: 3px 8px; border-radius: 3px; font-weight: bold; }}
    .grade-C {{ background: #ffc000; color: white; padding: 3px 8px; border-radius: 3px; font-weight: bold; }}
    .grade-D {{ background: #c00000; color: white; padding: 3px 8px; border-radius: 3px; font-weight: bold; }}
    .score {{ font-weight: bold; color: #1f4e78; }}
    .footer {{ margin-top: 40px; padding-top: 20px; border-top: 1px solid #d9d9d9; font-size: 12px; color: #888; }}
    .name {{ text-align: left; }}
    .small {{ font-size: 11px; color: #888; }}
</style>
</head>
<body>
    <h1>📊 主动基金月度筛选报告</h1>
    <p>报告日期: <strong>{report_date}</strong></p>

    <div class="summary-box">
        <strong>📌 本期摘要</strong><br>
        • 候选池规模: <strong>{candidate_count}</strong> 只(股票型 + 混合型 + QDII)<br>
        • 通过硬性筛选: <strong>{passed_count}</strong> 只<br>
        • 推荐 Top {top_n} 平均得分: <strong>{avg_score:.1f}</strong><br>
        • A 级数量: {a_count}; B 级数量: {b_count}
    </div>

    <h2>🏆 Top {top_n} 推荐列表</h2>
    <table>
        <thead>
            <tr>
                <th>排名</th>
                <th>基金代码</th>
                <th>基金简称</th>
                <th>基金经理</th>
                <th>综合得分</th>
                <th>评级</th>
                <th>入选关键原因 (维度×权重)</th>
            </tr>
        </thead>
        <tbody>
            {top_rows}
        </tbody>
    </table>

    <h2>📈 各项指标明细</h2>
    <table>
        <thead>
            <tr>
                <th>代码</th>
                <th>规模(亿)</th>
                <th>经理<br>任职(年)</th>
                <th>年化<br>收益(%)</th>
                <th>近3年<br>回撤(%)</th>
                <th>卡玛<br>比率</th>
                <th>波动率<br>(%)</th>
                <th>排名<br>分位</th>
                <th>熊市平均<br>回撤(%)</th>
                <th>行业<br>稳定性</th>
            </tr>
        </thead>
        <tbody>
            {detail_rows}
        </tbody>
    </table>

    <h2>🔍 各维度评分明细</h2>
    <table>
        <thead>
            <tr>
                <th>代码</th>
                <th>稳定性<br>(28%)</th>
                <th>熊市表现<br>(20%)</th>
                <th>任职<br>(15%)</th>
                <th>框架(卡玛)<br>(15%)</th>
                <th>风格(行业)<br>(12%)</th>
                <th>规模<br>(10%)</th>
                <th>综合</th>
            </tr>
        </thead>
        <tbody>
            {score_rows}
        </tbody>
    </table>

    {backtest_section}

    <div class="warning-box">
        <strong>⚠️ 重要提醒</strong><br>
        • 本报告基于历史数据回测,<strong>过去业绩不代表未来表现</strong>。<br>
        • 业绩稳定性 / 熊市表现 / 风格一致性均改为<strong>候选池内分位</strong>计算 (v2.1),避免顶部全员满分。<br>
        • 熊市区间在 <code>config.BEAR_MARKETS</code> 维护,月度脚本会自动检测新熊市段并在日志给出建议。<br>
        • 评分体系滚动回测 (5 起点 PIT) 用于验证选股能力,若长期 alpha 持平/为负请重新审视维度权重。<br>
        • 建议结合人工核查后再做投资决策,主动基金合理仓位不超过权益部分30%。
    </div>

    <div class="footer">
        本报告由自动化脚本生成 · 数据来源: 天天基金网 / AKShare · 评分卡 v2.1<br>
        如需调整筛选参数,请修改 src/config.py
    </div>
</body>
</html>
"""


def _grade_class(grade):
    if 'A' in grade:
        return 'A'
    if 'B' in grade:
        return 'B'
    if 'C' in grade:
        return 'C'
    if 'D' in grade:
        return 'D'
    return ''


def _fmt(val, fmt='{:.2f}', default='—'):
    """格式化数值,NaN 返回 default"""
    if val is None or pd.isna(val):
        return default
    try:
        return fmt.format(val)
    except (ValueError, TypeError):
        return str(val)


def _fmt_pct(v):
    """带正负号的百分点格式化, 用于超额收益等带符号场景"""
    if v is None or pd.isna(v):
        return '—'
    return f'{v:+.2f}'


def _build_backtest_html(backtest):
    """构造 HTML 邮件里的回测验证区块.
    支持两种输入: 单点 ({'summary': {...}}) / 滚动 ({'aggregate':..., 'per_window': [...]}).
    backtest 为 None 时返回空串."""
    if not backtest:
        return ''

    if 'aggregate' in backtest:
        return _build_rolling_backtest_html(backtest)

    if 'summary' not in backtest:
        return ''
    return _build_single_backtest_html(backtest['summary'])


def _build_rolling_backtest_html(rolling):
    agg = rolling.get('aggregate', {})
    per = rolling.get('per_window', [])

    def _v(d, key, fmt='{:.2f}'):
        val = d.get(key)
        if val is None or pd.isna(val):
            return '—'
        try:
            return fmt.format(val)
        except (ValueError, TypeError):
            return str(val)

    # 每窗口明细行
    detail_rows = []
    for s in per:
        detail_rows.append(f"""
            <tr>
                <td>{s.get('as_of', '—')}</td>
                <td>{s.get('hold_days', '—')}</td>
                <td class="score">{_v(s, 'top_n_avg_return_pct')}</td>
                <td>{_v(s, 'pool_passed_avg_return_pct')}</td>
                <td>{_v(s, 'universe_avg_return_pct')}</td>
                <td><strong>{_fmt_pct(s.get('excess_top_vs_pool_avg'))}</strong></td>
                <td>{_fmt_pct(s.get('excess_top_vs_universe_avg'))}</td>
                <td>{_v(s, 'win_rate_vs_pool_median_pct', '{:.0f}')}</td>
            </tr>
        """)

    avg_pool = agg.get('avg_top_alpha_vs_pool')
    pos_pool = agg.get('positive_alpha_windows_vs_pool', '—')
    diag = ''
    if isinstance(avg_pool, (int, float)) and not pd.isna(avg_pool):
        if avg_pool > 1:
            diag = '✅ 评分体系跨多窗口跑赢"硬筛池子"基准。'
        elif avg_pool > -1:
            diag = '⚖️ 评分体系与"硬筛池子"基准基本持平 — 当前模型对未来超额收益无显著解释力。'
        else:
            diag = '⚠️ 评分体系跨多窗口持续跑输"硬筛池子" — 需要重新审视评分维度的预测有效性。'

    return f"""
    <h2>🔬 评分体系滚动回测 (5 起点)</h2>
    <p class="small">用 {len(per)} 个历史时点的评分体系挑 Top, 持有至 <strong>{agg.get('hold_end', '—')}</strong>。
    每个起点都用当时的 PIT 数据重跑评分。</p>

    <div class="summary-box">
        <strong>📊 跨窗口聚合</strong><br>
        • 平均 Top 超额 vs 硬筛池子: <strong>{_fmt_pct(avg_pool)} pct</strong>;
        最差窗口: {_fmt_pct(agg.get('worst_top_alpha_vs_pool'))} pct<br>
        • 平均 Top 超额 vs 整宇宙: <strong>{_fmt_pct(agg.get('avg_top_alpha_vs_universe'))} pct</strong><br>
        • 正超额窗口数 (vs 池子): <strong>{pos_pool}</strong><br>
        • 平均 Top 胜率 vs 池子中位数: {_v(agg, 'avg_winrate_vs_pool_median', '{:.1f}')}%<br>
        <span class="small">{diag}</span>
    </div>

    <table>
        <thead>
            <tr>
                <th>起点 T</th><th>持有天</th>
                <th>Top 平均(%)</th><th>池子平均(%)</th><th>宇宙平均(%)</th>
                <th>vs 池子</th><th>vs 宇宙</th><th>胜率(vs 池中位)</th>
            </tr>
        </thead>
        <tbody>
            {''.join(detail_rows)}
        </tbody>
    </table>
    <p class="small">说明: 每窗口用 as_of_date 时点的 PIT 数据 (经理任职 / 收益 / 回撤 / 卡玛全部按当时数据重算)。
    宇宙存活者偏差不可避免; 规模/在管基金数用当前值代理 (已知偏差)。</p>
    """


def _build_single_backtest_html(s):
    """兼容旧版单点回测格式"""
    def _v(key, fmt='{:.2f}'):
        val = s.get(key)
        if val is None or pd.isna(val):
            return '—'
        try:
            return fmt.format(val)
        except (ValueError, TypeError):
            return str(val)

    return f"""
    <h2>🔬 评分体系回测验证</h2>
    <p class="small">PIT 回测: 用 <strong>{s.get('as_of', '—')}</strong> 时点的评分体系挑 Top {s.get('top_n_size', '—')},
    持有至 <strong>{s.get('hold_end', '—')}</strong> ({s.get('hold_days', '—')} 天)。</p>
    <table>
        <thead><tr><th>口径</th><th>Top {s.get('top_n_size', 'N')}</th><th>硬筛池</th><th>整宇宙</th></tr></thead>
        <tbody>
            <tr><td>平均持有期收益(%)</td><td class="score">{_v('top_n_avg_return_pct')}</td><td>{_v('pool_passed_avg_return_pct')}</td><td>{_v('universe_avg_return_pct')}</td></tr>
            <tr><td>Top 超额(pct)</td><td colspan="2">vs 池子: <strong>{_fmt_pct(s.get('excess_top_vs_pool_avg'))}</strong></td><td>vs 宇宙: <strong>{_fmt_pct(s.get('excess_top_vs_universe_avg'))}</strong></td></tr>
        </tbody>
    </table>
    """


def _fmt_bear_dd(row):
    """熊市平均回撤显示: NaN 且熊市数=0 时显示'未经历熊市',否则显示'—'"""
    dd = row.get('熊市平均回撤')
    if pd.notna(dd):
        return f'{dd:.1f}'
    bc = row.get('熊市数', 0)
    if pd.isna(bc):
        bc = 0
    return '未经历熊市' if int(bc) == 0 else '—'


def generate_html_report(top_n_df, all_df, backtest=None):
    """生成 HTML 邮件正文. backtest: src.backtest.run_backtest 返回值, 可选"""
    if len(top_n_df) == 0:
        return _generate_empty_report(all_df)

    candidate_count = len(all_df)
    passed_count = int(all_df['硬筛通过'].sum()) if '硬筛通过' in all_df.columns else 0
    avg_score = top_n_df['综合得分'].mean()
    a_count = int((top_n_df['评级'] == 'A 级').sum())
    b_count = int((top_n_df['评级'] == 'B 级').sum())

    # Top 排名表
    top_rows = []
    for i, (_, row) in enumerate(top_n_df.iterrows(), 1):
        gc = _grade_class(row['评级'])
        manager = row.get('基金经理', '')
        if pd.isna(manager):
            manager = ''
        reason = row.get('入选原因', '') or ''
        top_rows.append(f"""
            <tr>
                <td class="rank">{i}</td>
                <td>{row['基金代码']}</td>
                <td class="name">{row['基金简称']}</td>
                <td>{manager}</td>
                <td class="score">{row['综合得分']:.1f}</td>
                <td><span class="grade-{gc}">{row['评级']}</span></td>
                <td class="small name">{reason}</td>
            </tr>
        """)

    # 指标明细表
    detail_rows = []
    for _, row in top_n_df.iterrows():
        detail_rows.append(f"""
            <tr>
                <td>{row['基金代码']}</td>
                <td>{_fmt(row.get('基金规模'), '{:.1f}')}</td>
                <td>{_fmt(row.get('经理任职年限'), '{:.1f}')}</td>
                <td>{_fmt(row.get('年化收益率'), '{:.2f}')}</td>
                <td>{_fmt(row.get('近3年最大回撤'), '{:.2f}')}</td>
                <td>{_fmt(row.get('卡玛比率'), '{:.2f}')}</td>
                <td>{_fmt(row.get('年化波动率'), '{:.1f}')}</td>
                <td>{_fmt(row.get('业绩排名分位'), '{:.0f}')}</td>
                <td>{_fmt_bear_dd(row)}</td>
                <td>{_fmt(row.get('行业稳定性'), '{:.2f}')}</td>
            </tr>
        """)

    # 评分明细表
    score_rows = []
    for _, row in top_n_df.iterrows():
        score_rows.append(f"""
            <tr>
                <td>{row['基金代码']}</td>
                <td>{_fmt(row.get('得分_稳定性'), '{:.0f}')}</td>
                <td>{_fmt(row.get('得分_熊市'), '{:.0f}')}</td>
                <td>{_fmt(row.get('得分_任职'), '{:.0f}')}</td>
                <td>{_fmt(row.get('得分_框架'), '{:.0f}')}</td>
                <td>{_fmt(row.get('得分_风格'), '{:.0f}')}</td>
                <td>{_fmt(row.get('得分_规模'), '{:.0f}')}</td>
                <td class="score">{row['综合得分']:.1f}</td>
            </tr>
        """)

    html = HTML_TEMPLATE.format(
        report_date=datetime.now().strftime('%Y-%m-%d'),
        candidate_count=candidate_count,
        passed_count=passed_count,
        top_n=len(top_n_df),
        avg_score=avg_score,
        a_count=a_count,
        b_count=b_count,
        top_rows=''.join(top_rows),
        detail_rows=''.join(detail_rows),
        score_rows=''.join(score_rows),
        backtest_section=_build_backtest_html(backtest),
    )
    return html


def _generate_empty_report(all_df):
    """没有基金通过筛选时的报告"""
    return f"""
    <html><body style="font-family: 'Microsoft YaHei'; padding: 20px;">
        <h2>📊 主动基金月度筛选报告</h2>
        <p>报告日期: {datetime.now().strftime('%Y-%m-%d')}</p>
        <div style="background: #ffe4e1; padding: 15px; border-left: 4px solid #c00;">
            <strong>⚠️ 本期未有基金通过筛选</strong><br>
            候选池规模: {len(all_df)}<br>
            建议检查数据源是否正常,或调整 config.py 的筛选阈值。
        </div>
    </body></html>
    """


# ============================================================================
# Excel 详细报告
# ============================================================================
def generate_excel_report(top_n_df, all_df, output_path, backtest=None):
    """生成 Excel 详细附件. backtest: src.backtest.run_backtest 返回值, 可选,
    存在时新增 "回测验证" sheet"""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws_top = wb.active
    ws_top.title = 'Top推荐'

    # 字体样式
    F_TITLE = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
    F_HEADER = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    F_BODY = Font(name='微软雅黑', size=10)
    FILL_TITLE = PatternFill('solid', start_color='1F4E78')
    FILL_HEADER = PatternFill('solid', start_color='2E75B6')
    ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='BFBFBF')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ========== Sheet 1: Top 推荐 ==========
    top_cols = [
        '排名', '基金代码', '基金简称', '基金类型', '基金经理',
        '基金规模', '经理任职年限', '年化收益率', '近3年最大回撤',
        '卡玛比率', '夏普比率', '年化波动率', '业绩排名分位',
        '熊市数', '熊市平均回撤', '行业稳定性',
        '得分_稳定性', '得分_熊市', '得分_任职',
        '得分_框架', '得分_风格', '得分_规模',
        '综合得分', '评级', '入选原因',
    ]

    last_col_letter = get_column_letter(len(top_cols))
    ws_top.merge_cells(f'A1:{last_col_letter}1')
    ws_top['A1'] = f"主动基金月度筛选报告 · Top {len(top_n_df)} 推荐"
    ws_top['A1'].font = F_TITLE
    ws_top['A1'].fill = FILL_TITLE
    ws_top['A1'].alignment = ALIGN_C
    ws_top.row_dimensions[1].height = 28

    ws_top.merge_cells(f'A2:{last_col_letter}2')
    ws_top['A2'] = f"报告日期: {datetime.now().strftime('%Y-%m-%d')}    评分卡 v2.1"
    ws_top['A2'].font = Font(name='微软雅黑', size=10, italic=True, color='595959')
    ws_top['A2'].alignment = ALIGN_L

    # 表头
    for j, col in enumerate(top_cols, 1):
        c = ws_top.cell(row=4, column=j, value=col)
        c.font = F_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_C
        c.border = BORDER

    # 数据行
    for i, (_, row) in enumerate(top_n_df.iterrows(), start=5):
        ws_top.cell(row=i, column=1, value=i - 4)
        for j, col in enumerate(top_cols[1:], 2):
            val = row.get(col, '')
            # 熊市平均回撤 NaN 时给出温柔提示
            if col == '熊市平均回撤' and pd.isna(val):
                bc = row.get('熊市数', 0)
                if pd.isna(bc):
                    bc = 0
                val = '未经历熊市' if int(bc) == 0 else ''
            elif pd.isna(val):
                val = ''
            cell = ws_top.cell(row=i, column=j, value=val)
            cell.font = F_BODY
            cell.alignment = ALIGN_C
            cell.border = BORDER
            # 数值格式
            if col in ('基金规模', '经理任职年限', '年化波动率', '熊市平均回撤'):
                cell.number_format = '0.00'
            elif col in ('年化收益率', '近3年最大回撤'):
                cell.number_format = '0.00'
            elif col in ('夏普比率', '卡玛比率', '行业稳定性'):
                cell.number_format = '0.00'
            elif col == '综合得分':
                cell.number_format = '0.0'
            elif col.startswith('得分_'):
                cell.number_format = '0'

    # 列宽
    widths = {'A': 6, 'B': 10, 'C': 25, 'D': 12, 'E': 12}
    for col, w in widths.items():
        ws_top.column_dimensions[col].width = w
    for j in range(6, len(top_cols) + 1):
        ws_top.column_dimensions[get_column_letter(j)].width = 11

    # 冻结
    ws_top.freeze_panes = 'D5'

    # 综合得分色阶
    if len(top_n_df) > 0:
        score_col = top_cols.index('综合得分') + 1
        col_letter = get_column_letter(score_col)
        ws_top.conditional_formatting.add(
            f'{col_letter}5:{col_letter}{4 + len(top_n_df)}',
            ColorScaleRule(start_type='num', start_value=60, start_color='F8696B',
                           mid_type='num', mid_value=75, mid_color='FFEB84',
                           end_type='num', end_value=95, end_color='63BE7B')
        )

    # ========== Sheet 2: 完整候选池 ==========
    ws_all = wb.create_sheet('完整候选池')
    all_cols = [c for c in top_cols if c != '排名'] + ['硬筛通过', '淘汰原因']
    all_cols = [c for c in all_cols if c in all_df.columns]

    ws_all.merge_cells(f'A1:{get_column_letter(len(all_cols))}1')
    ws_all['A1'] = f"完整候选池分析(共 {len(all_df)} 只)"
    ws_all['A1'].font = F_TITLE
    ws_all['A1'].fill = FILL_TITLE
    ws_all['A1'].alignment = ALIGN_C
    ws_all.row_dimensions[1].height = 28

    for j, col in enumerate(all_cols, 1):
        c = ws_all.cell(row=3, column=j, value=col)
        c.font = F_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_C
        c.border = BORDER

    # 按综合得分降序输出
    sorted_all = all_df.sort_values('综合得分', ascending=False) if '综合得分' in all_df.columns else all_df
    for i, (_, row) in enumerate(sorted_all.iterrows(), start=4):
        for j, col in enumerate(all_cols, 1):
            val = row.get(col, '')
            if col == '熊市平均回撤' and pd.isna(val):
                bc = row.get('熊市数', 0)
                if pd.isna(bc):
                    bc = 0
                val = '未经历熊市' if int(bc) == 0 else ''
            elif pd.isna(val):
                val = ''
            cell = ws_all.cell(row=i, column=j, value=val)
            cell.font = F_BODY
            cell.alignment = ALIGN_C
            cell.border = BORDER

    for j in range(1, len(all_cols) + 1):
        ws_all.column_dimensions[get_column_letter(j)].width = 12
    ws_all.freeze_panes = 'D4'

    # ========== Sheet 3: 评分规则说明 ==========
    ws_rule = wb.create_sheet('评分规则')
    ws_rule.merge_cells('A1:D1')
    ws_rule['A1'] = '评分规则说明'
    ws_rule['A1'].font = F_TITLE
    ws_rule['A1'].fill = FILL_TITLE
    ws_rule['A1'].alignment = ALIGN_C
    ws_rule.row_dimensions[1].height = 28

    rules = [
        ('硬性筛选(6项,NaN 一律放行)', '', '', ''),
        ('1. 经理任职年限', '≥ 3 年', '放宽至3年纳入近年新锐', '经理表'),
        ('2. 基金成立时间', '≥ 3 年', '新基金数据不可靠', '排名表'),
        ('3. 基金规模', '2 - 100 亿', '过大调仓难,过小有清盘风险', '排名表'),
        ('4. 经理在管基金数', '≤ 5 只', '管太多说明在挂名', '经理表'),
        ('5. 近3年最大回撤', '≤ 45%(绝对)', '风控能力检验,改绝对避免候选池偏差', '净值计算'),
        ('6. 近1年收益兜底', '≥ -25%', '过滤长期好但近期暴雷的基金', '排名表'),
        ('', '', '', ''),
        ('软性评分(6维度,加权满分100,NaN→50)', '权重', '评分逻辑 (v2.1 候选池内分位)', '指标'),
        ('1. 业绩稳定性', '28%', '近3年收益候选池内分位; ≤5→100/≤10→92/≤20→82/≤35→68/≤55→55/≤75→40/≤90→28/其余 18', '池内分位'),
        ('2. 熊市相对表现', '20%', '历轮熊市平均回撤池内分位; ≤8→100/≤18→90/≤35→75/≤55→60/≤75→42/其余 28 + (熊市数-1)×2', '净值×熊市区间'),
        ('3. 经理任职年限', '15%', '分段线性: 3年→55, 5→72, 7→85, 10→95, ≥15→100 (边际递减)', '年限分档'),
        ('4. 投资框架', '15%', '卡玛比率(年化收益/最大回撤),≥0.8 满分', '净值计算'),
        ('5. 风格一致性', '12%', '行业相似度池内分位; ≤10→100/≤25→88/≤45→72/≤65→55/≤85→40/其余 22 (数据不足回退波动率)', '行业配置历史'),
        ('6. 规模适中度', '10%', '钟形分段: [5,30]亿满分; 2→60, 50→85, 100→55, ≥200→30', '规模分档'),
        ('', '', '', ''),
        ('熊市区间 (config.BEAR_MARKETS 可手工维护)', '', '月度运行会基于候选池中位 NAV 自动检测新熊市段并日志建议', ''),
        ('', '', '', ''),
        ('评级标准', '', '', ''),
        ('A 级', '≥ 85', '强烈推荐,核心持仓', '权益部分10-20%'),
        ('B 级', '70 - 85', '推荐,可作备选', '权益部分5-10%'),
        ('C 级', '60 - 70', '观望,继续跟踪', '不持有'),
        ('D 级', '< 60', '不推荐', '不持有'),
    ]
    headers = ['项目', '标准/权重', '说明', '数据来源']
    for j, h in enumerate(headers, 1):
        c = ws_rule.cell(row=3, column=j, value=h)
        c.font = F_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_C
        c.border = BORDER

    for i, row_data in enumerate(rules, start=4):
        for j, val in enumerate(row_data, 1):
            c = ws_rule.cell(row=i, column=j, value=val)
            c.font = F_BODY
            c.alignment = Alignment(horizontal='left' if j == 1 else 'center', vertical='center', wrap_text=True)
            c.border = BORDER
            # 子标题加粗
            if row_data[0] in ('硬性筛选(6项,NaN 一律放行)', '软性评分(6维度,加权满分100,NaN→50)', '评级标准'):
                c.font = Font(name='微软雅黑', size=11, bold=True, color='1F4E78')
                c.fill = PatternFill('solid', start_color='D9E1F2')

    ws_rule.column_dimensions['A'].width = 28
    ws_rule.column_dimensions['B'].width = 22
    ws_rule.column_dimensions['C'].width = 35
    ws_rule.column_dimensions['D'].width = 22

    # ========== Sheet 4: 回测验证 (仅当传入 backtest 时) ==========
    if backtest:
        if 'aggregate' in backtest:
            _write_rolling_backtest_sheet(wb, backtest, F_TITLE, F_HEADER, F_BODY,
                                          FILL_TITLE, FILL_HEADER, ALIGN_C, ALIGN_L, BORDER)
        elif 'summary' in backtest:
            _write_single_backtest_sheet(wb, backtest, F_TITLE, F_HEADER, F_BODY,
                                         FILL_TITLE, FILL_HEADER, ALIGN_C, ALIGN_L, BORDER)

    wb.save(output_path)
    logger.info(f'Excel 报告已保存: {output_path}')
    return output_path


def _write_rolling_backtest_sheet(wb, rolling, F_TITLE, F_HEADER, F_BODY,
                                  FILL_TITLE, FILL_HEADER, ALIGN_C, ALIGN_L, BORDER):
    """滚动回测 sheet: 聚合 + 每窗口明细 + 每窗口 Top N"""
    ws = wb.create_sheet('回测验证')
    agg = rolling.get('aggregate', {})
    per = rolling.get('per_window', [])

    ws.merge_cells('A1:H1')
    ws['A1'] = f"评分体系滚动回测 ({agg.get('windows', '—')} 个起点 PIT)"
    ws['A1'].font = F_TITLE
    ws['A1'].fill = FILL_TITLE
    ws['A1'].alignment = ALIGN_C
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:H2')
    ws['A2'] = f"持有至: {agg.get('hold_end', '—')} | 起点: {', '.join(agg.get('as_of_dates') or []) or '—'}"
    ws['A2'].font = Font(name='微软雅黑', size=10, italic=True, color='595959')
    ws['A2'].alignment = ALIGN_L

    # 聚合区
    agg_rows = [
        ('指标', '值'),
        ('窗口数', agg.get('windows')),
        ('平均 Top 超额 vs 池子 (pct)', agg.get('avg_top_alpha_vs_pool')),
        ('平均 Top 超额 vs 宇宙 (pct)', agg.get('avg_top_alpha_vs_universe')),
        ('最差 Top 超额 vs 池子 (pct)', agg.get('worst_top_alpha_vs_pool')),
        ('最差 Top 超额 vs 宇宙 (pct)', agg.get('worst_top_alpha_vs_universe')),
        ('正超额窗口数 (vs 池子)', agg.get('positive_alpha_windows_vs_pool')),
        ('正超额窗口数 (vs 宇宙)', agg.get('positive_alpha_windows_vs_universe')),
        ('平均 Top 胜率 vs 池子中位数 (%)', agg.get('avg_winrate_vs_pool_median')),
        ('平均 Top 胜率 vs 宇宙中位数 (%)', agg.get('avg_winrate_vs_universe_median')),
    ]
    for i, row_data in enumerate(agg_rows, start=4):
        for j, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=j, value=val if val is not None else '')
            c.font = F_HEADER if i == 4 else F_BODY
            if i == 4:
                c.fill = FILL_HEADER
            c.alignment = ALIGN_C if j != 1 else ALIGN_L
            c.border = BORDER
            if i > 4 and j == 2 and isinstance(val, (int, float)):
                c.number_format = '0.00'

    # 每窗口明细
    start = 4 + len(agg_rows) + 2
    ws.merge_cells(f'A{start}:H{start}')
    title_cell = ws.cell(row=start, column=1, value='每窗口明细')
    title_cell.font = Font(name='微软雅黑', size=11, bold=True, color='1F4E78')
    title_cell.fill = PatternFill('solid', start_color='D9E1F2')
    title_cell.alignment = ALIGN_C

    headers = ['起点 T', '持有天', 'Top 平均(%)', '池子平均(%)', '宇宙平均(%)',
               'Top vs 池子(pct)', 'Top vs 宇宙(pct)', '胜率 vs 池中位(%)']
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=start + 1, column=j, value=h)
        c.font = F_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_C
        c.border = BORDER

    for i, s in enumerate(per, start=1):
        r = start + 1 + i
        vals = [
            s.get('as_of'), s.get('hold_days'),
            s.get('top_n_avg_return_pct'), s.get('pool_passed_avg_return_pct'),
            s.get('universe_avg_return_pct'),
            s.get('excess_top_vs_pool_avg'), s.get('excess_top_vs_universe_avg'),
            s.get('win_rate_vs_pool_median_pct'),
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=j, value=v if v is not None else '')
            c.font = F_BODY
            c.alignment = ALIGN_C
            c.border = BORDER
            if isinstance(v, (int, float)) and j > 1:
                c.number_format = '0.00'

    widths = [14, 8, 12, 13, 13, 14, 14, 16]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _write_single_backtest_sheet(wb, backtest, F_TITLE, F_HEADER, F_BODY,
                                 FILL_TITLE, FILL_HEADER, ALIGN_C, ALIGN_L, BORDER):
    """兼容旧版单点回测 sheet"""
    ws = wb.create_sheet('回测验证')
    s = backtest['summary']

    ws.merge_cells('A1:D1')
    ws['A1'] = '评分体系回测验证 (PIT)'
    ws['A1'].font = F_TITLE
    ws['A1'].fill = FILL_TITLE
    ws['A1'].alignment = ALIGN_C
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:D2')
    ws['A2'] = (
        f"回测起点 T = {s.get('as_of', '—')}, 持有至 {s.get('hold_end', '—')} "
        f"({s.get('hold_days', '—')} 天), "
        f"宇宙 {s.get('pool_universe_size', '—')} 只, "
        f"硬筛通过 {s.get('pool_passed_hard_filter', '—')} 只"
    )
    ws['A2'].font = Font(name='微软雅黑', size=10, italic=True, color='595959')
    ws['A2'].alignment = ALIGN_L

    bt_rows = [
        ('指标', 'Top N', '通过硬筛池子', '整宇宙'),
        ('平均持有期收益 (%)',
            s.get('top_n_avg_return_pct'),
            s.get('pool_passed_avg_return_pct'),
            s.get('universe_avg_return_pct')),
        ('中位数 (%)',
            s.get('top_n_median_return_pct'),
            s.get('pool_passed_median_return_pct'),
            s.get('universe_median_return_pct')),
        ('Top 超额 (pct)',
            None,
            f"vs 池子: {s.get('excess_top_vs_pool_avg', '—')}",
            f"vs 宇宙: {s.get('excess_top_vs_universe_avg', '—')}"),
        ('Top 胜率 vs 中位数 (%)',
            None,
            f"vs 池子: {s.get('win_rate_vs_pool_median_pct', '—')}",
            f"vs 宇宙: {s.get('win_rate_vs_universe_median_pct', '—')}"),
    ]
    for i, row_data in enumerate(bt_rows, start=4):
        for j, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=j, value=val if val is not None else '')
            c.font = F_HEADER if i == 4 else F_BODY
            if i == 4:
                c.fill = FILL_HEADER
            c.alignment = ALIGN_C
            c.border = BORDER
            if i > 4 and j > 1 and isinstance(val, (int, float)):
                c.number_format = '0.00'

    top_bt = backtest.get('top_n')
    if top_bt is not None and len(top_bt) > 0:
        start_row = 4 + len(bt_rows) + 2
        ws.merge_cells(f'A{start_row}:D{start_row}')
        ws.cell(row=start_row, column=1, value='Top N 持有期收益明细').font = Font(
            name='微软雅黑', size=11, bold=True, color='1F4E78')
        ws.cell(row=start_row, column=1).fill = PatternFill('solid', start_color='D9E1F2')
        ws.cell(row=start_row, column=1).alignment = ALIGN_C

        headers_bt = ['排名', '基金代码', '基金简称', '持有期收益 (%)']
        for j, h in enumerate(headers_bt, 1):
            c = ws.cell(row=start_row + 1, column=j, value=h)
            c.font = F_HEADER
            c.fill = FILL_HEADER
            c.alignment = ALIGN_C
            c.border = BORDER

        for i, (_, r) in enumerate(top_bt.iterrows(), start=1):
            ws.cell(row=start_row + 1 + i, column=1, value=i)
            ws.cell(row=start_row + 1 + i, column=2, value=r.get('基金代码', ''))
            ws.cell(row=start_row + 1 + i, column=3, value=r.get('基金简称', ''))
            ret = r.get('持有期收益')
            ws.cell(row=start_row + 1 + i, column=4,
                    value=float(ret) if pd.notna(ret) else '')
            for j in range(1, 5):
                cell = ws.cell(row=start_row + 1 + i, column=j)
                cell.font = F_BODY
                cell.alignment = ALIGN_C if j != 3 else ALIGN_L
                cell.border = BORDER
                if j == 4 and pd.notna(ret):
                    cell.number_format = '0.00'

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 22
